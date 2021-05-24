import csv

import pandas as pd
from openpyxl import load_workbook

from utils import strip_accents_spain


URL_UBIGEO_INEI = "https://raw.githubusercontent.com/rmaceissoft/ubigeo-peru/fix/update_inei_ubigeos/equivalencia-ubigeos-oti-concytec.csv"


def main():
    filename = "./data/distritos_ceplan_2019.xlsx"
    wb = load_workbook(filename)
    sheet = wb["Sheet1"]
    # load csv with ubigeo into dataframme
    ubigeo_df = pd.read_csv(
        URL_UBIGEO_INEI,
        encoding="latin",
        dtype={"cod_dep_inei": str},
        usecols=[
            "desc_dep_inei",
            "desc_prov_inei",
            "desc_ubigeo_inei",
            "cod_ubigeo_inei",
        ],
    )
    with open("./data/distritos_peru.csv", "w", newline="") as output_csv:
        output_writer = csv.writer(output_csv)
        output_writer.writerow(
            ["departamento", "provincia", "distrito", "latitud", "longitud", "ubigeo"]
        )
        for row in sheet.iter_rows(min_row=2):
            district_ubigeo = row[1].value
            district_latitude = row[13].value
            district_longitude = row[14].value
            print(
                f"extracting district: ubigeo={district_ubigeo}; lat={district_latitude}; lon={district_longitude}"
            )
            result_df = ubigeo_df[ubigeo_df["cod_ubigeo_inei"] == district_ubigeo]
            if result_df.empty:
                print(f"WARNING: {district_ubigeo} was not found at ubigeo database")
            else:
                department = strip_accents_spain(result_df.values[0][0])
                province = strip_accents_spain(result_df.values[0][1])
                district = strip_accents_spain(result_df.values[0][3])
                output_writer.writerow(
                    [
                        department,
                        province,
                        district,
                        district_latitude,
                        district_longitude,
                        district_ubigeo,
                    ]
                )


if __name__ == "__main__":
    main()
